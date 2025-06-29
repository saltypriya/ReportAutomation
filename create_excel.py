import openpyxl
from openpyxl.styles import Font

def create_claims_excel():
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claims Data"
    
    # Define column headers
    headers = [
        "insured_name",
        "address",
        "insurer",
        "claim_number",
        "date_of_inspection",
        "date_of_loss",
        "date_of_report",
        "type_of_loss",
        "cause_of_loss",
        "indemnity_work",
        "listing_pricing_reserve",
        "contents_loss_reserve"
    ]
    
    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Sample claim data
    claims_data = [
        [
            "ABCD Liu",
            "B-63 White Heather Blvd, Scarborough, ON M1V 1P6",
            "ABC Insurance",
            "PR1234",
            "2025-03-31",
            "2025-03-28",
            "2025-04-01",
            "Fire",
            "Fire originated at neighboring house (61 White Heather Blvd)",
            20000,
            4000,
            4000
        ],
        [
            "Michael Chen",
            "112 Pine Road, Markham, ON L3R 2G5",
            "XYZ Insurance",
            "CL7890",
            "2025-04-15",
            "2025-04-12",
            "2025-04-16",
            "Water Damage",
            "Burst pipe in upstairs bathroom",
            25000,
            5000,
            7500
        ],
        [
            "Sarah Johnson",
            "45 Oak Avenue, Toronto, ON M4B 1B2",
            "DEF Insurance",
            "CL2468",
            "2025-05-10",
            "2025-05-08",
            "2025-05-11",
            "Theft",
            "Break-in through rear window",
            18000,
            3500,
            6200
        ],
        [
            "David Wilson",
            "88 Elm Street, Mississauga, ON L5M 3H2",
            "GHI Insurance",
            "CL1357",
            "2025-06-05",
            "2025-06-01",
            "2025-06-06",
            "Storm Damage",
            "Fallen tree damaged roof",
            30000,
            6000,
            8500
        ],
        [
            "Emily Brown",
            "27 Maple Lane, Brampton, ON L6Y 4T2",
            "JKL Insurance",
            "CL3691",
            "2025-07-20",
            "2025-07-18",
            "2025-07-21",
            "Smoke Damage",
            "Kitchen fire in adjacent unit",
            22000,
            4500,
            5800
        ]
    ]
    
    # Add data rows
    for row_num, claim in enumerate(claims_data, 2):
        for col_num, value in enumerate(claim, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save("claim_data.xlsx")
    print("Excel file 'claim_data.xlsx' created successfully!")

if __name__ == "__main__":
    create_claims_excel()